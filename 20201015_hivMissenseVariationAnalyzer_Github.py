inputAlignment = "HIV1_ALL_2017_nef_PRO.fasta" # An alignment file produced by AliView.
inputAlignmentPath = "/Users/newumuser/Desktop/Nef Polymorphism Calculator/"

# Dependencies.
###############
from openpyxl import Workbook

# Get the file information. 
###########################
input = open(inputAlignmentPath + inputAlignment, "r")
lines = input.readlines() # The readline() function returns a list of lines in the file.
input.close()

# Make a dictionary where the Fasta sequence name is the key, and the Fasta sequence is the value.
##################################################################################################
sequences = {}
i = 0
while i < len(lines):
	line = lines[i]
	if ">" in line:
		line = line.split()[0] # split removes newline at the end of the line.
		sequences[line] = lines[i+1] 
	i += 1

##### Remove sequences with ambiguous characters
i = 0
j = 0
a = 0
b = 0
z = 0
newlyAlignedSequences = {}
removedSequences = []
for seq in sequences:
	# if "*" in sequences[seq][:-1]:
	# 	if seq not in removedSequences:
	# 		removedSequences.append(seq)
	# 		z += 1
	if "X" in sequences[seq]:
		if seq not in removedSequences:
			removedSequences.append(seq)
			i += 1
	elif "#" in sequences[seq]:
		if seq not in removedSequences:
			removedSequences.append(seq)
			j += 1
	elif "?" in sequences[seq]:
		if seq not in removedSequences:
			removedSequences.append(seq)
			a += 1
	elif "$" in sequences[seq]:
		if seq not in removedSequences:
			removedSequences.append(seq)
			b += 1
	else:
		newlyAlignedSequences[seq] = sequences[seq]

print("There are", len(removedSequences), "sequences excluded from the Envelope Fasta because of unusual characters.")
print("There are", len(newlyAlignedSequences), "sequences in the modified Envelope Fasta.")

# This following algorithm only works if all the sequences are the same length, and they should be when they come out of AliView.
# Get the sequence length.
##########################
keys = newlyAlignedSequences.keys()
keysList = list(keys)
firstKeyInSequences = keysList[0] 
sequenceLength = len(newlyAlignedSequences[firstKeyInSequences])

# Make a dictionary of sequence positions using the variable sequenceLength.
############################################################################
aminoAcidOccurrencePerPositionDict = {}
i = 1
while i < sequenceLength:
	aminoAcidOccurrencePerPositionDict[i] = {}
	i += 1

# Collect amino acid occurences per position in the dictinonary of dictionaries, variantDict.
#############################################################################################
for key in keys:
	# key is a sequence identifier, and value is a sequence
	sequence = newlyAlignedSequences[key]
	i = 0
	while i < len(sequence) - 1:
		positionDict = aminoAcidOccurrencePerPositionDict[i+1]
		aminoAcid = sequence[i]
		if aminoAcid not in positionDict:
			positionDict[aminoAcid] = [key,]
		else:
			positionDict[aminoAcid].append(key)
		i += 1

# Make a dictionary of amino acid occurrence frequences per position.
#####################################################################
aminoAcidFrequencyDict = {}
positionKeys = aminoAcidOccurrencePerPositionDict.keys()
for positionKey in positionKeys:
	aminoAcidKeys = aminoAcidOccurrencePerPositionDict[positionKey].keys()
	frequency = {}
	for aminoAcidKey in aminoAcidKeys:
		f = 100.*(len(aminoAcidOccurrencePerPositionDict[positionKey][aminoAcidKey])/float(len(newlyAlignedSequences))) # Expressed as percent of number of sequences.
		frequency[aminoAcidKey] = f  
	aminoAcidFrequencyDict[positionKey] = frequency

alignedVA47081Nef = 'MGGKWSKCS----M-GWPTVRERIRQVEP------------------AAE--PAAAGVGAVSQDLEQRGAITSSN-T--NNAACAWLEAQ----EEEEVGFPVRPQVPLRPMTYK-GAL-DLSHFLK-EKGGLEGLVWCQ-RRQDI---LDLWVYNTQGYF-PDWQNYTPGPGIR-YPLTFGWCFKLVPVDKDQVEKE-NEGEDKNLLSP-MSLHG-MEDTEK-----------EVLVWKFDSRLAFHHMARELHPEYY---------KDC*--------------------'

# Rank the amino acid variants by occurence at each sequence position.
######################################################################
rankedAminoAcidVariations = {}
for positionKey in positionKeys:
	aminoAcidKeys = aminoAcidOccurrencePerPositionDict[positionKey].keys()
	for aminoAcidKey in aminoAcidKeys:
		for fastaSequenceIdentifier in aminoAcidOccurrencePerPositionDict[positionKey][aminoAcidKey]:
			rankedAminoAcidVariations[(aminoAcidFrequencyDict[positionKey][aminoAcidKey], aminoAcidKey, positionKey, fastaSequenceIdentifier)] = None
keys = rankedAminoAcidVariations.keys()
keysSorted = sorted(keys)

for key in keysSorted:
	print(key)

# Open two new Excel workbook. 
# One for missense mutations below the frequency threshod
# One for missense mutations above the frequency threshold
##########################################################
book = Workbook()
sheet = book.active
sheet.title = "HIV rare variations"
ws3 = book.create_sheet(title = "occurrence")
ws4 = book.create_sheet(title = "average rare occurrences")

cell = sheet.cell(row=1, column=1)
cell.value = "%" + " Frequency"
cell = sheet.cell(row=1, column=2)
cell.value = "Amino Acid"
cell = sheet.cell(row=1, column=3)
cell.value = "Aligned Position" 
cell = sheet.cell(row=1, column=4)
cell.value = "Fasta ID" 

cell3 = ws3.cell(row=1, column=1)
cell3.value = "FASTA ID"
cell3 = ws3.cell(row=1, column=2)
cell3.value = "Number of rare mutations"

cell4 = ws4.cell(row=1, column=1)
cell4.value = "Number of rare polymorphisms"
cell4 = ws4.cell(row=1, column=2)
cell4.value = "Number of sequences"

r = 2
z = 2
rareSequenceList = []
# Write variants below the frequency cutoff to one Excel workbook.
# Write variants above the frequency cutoff to another Excel workbook.
#######################################
for key in keysSorted:
	if key[0] <= 2.5:
		cell = sheet.cell(row=r, column=1)
		cell.value = key[0]
		cell = sheet.cell(row=r, column=2)
		cell.value = key[1]
		cell = sheet.cell(row=r, column=3)
		cell.value = key[2]
		cell = sheet.cell(row=r, column=4)
		cell.value = key[3]
		rareSequenceList.append(key[3])
		r += 1
	# else: 
	# 	cell2 = sheet2.cell(row=z, column=1)
	# 	cell2.value = key[0]
	# 	cell2 = sheet2.cell(row=z, column=2)
	# 	cell2.value = key[1]
	# 	cell2 = sheet2.cell(row=z, column=3)
	# 	cell2.value = key[2]
	# 	cell2 = sheet2.cell(row=z, column=4)
	# 	cell2.value = key[3]
	# 	z += 1
print(rareSequenceList)

# Determine the number of rare mutations in each sequence
#########################################################
fakeList = []
numberList = []
g = 2
for item in rareSequenceList:
	print(rareSequenceList)
	
	if item in fakeList:
		pass
	else:
		fakeList.append(item)
		number = rareSequenceList.count(item)
		numberList.append(number)
		cell3 = ws3.cell(row=g, column=1)
		cell3.value = item
		cell3 = ws3.cell(row=g, column=2)
		cell3.value = number
		g += 1

# Determine the number of sequences with a certain number of rare mutations
###########################################################################
fakerList = []
h = 2
for number in numberList:
	print(numberList)
	if number in fakerList:
		pass
	else:
		fakerList.append(number)
		occurrence = numberList.count(number)
		cell4 = ws4.cell(row=h, column=1)
		cell4.value = number
		cell4 = ws4.cell(row=h, column=2)
		cell4.value = occurrence
		h += 1
	
# Save the Excel workbook.
#########################
book.save(inputAlignmentPath + inputAlignment.split(".fasta")[0] + "-below" + cutoff + "Variations" + ".xlsx")






